#!/usr/bin/env python3
"""
exporter.py — Part 0 one-time migration (HTML COUNTRY_DATA -> Sheets layout).

Takes the COUNTRY_DATA object (extracted from the HTML, saved as JSON) and
writes it into the flat, section-based sheet layout defined in
DATA_PIPELINE_ARCHITECTURE.md Part 1. This is the mirror image of converter.py.

Two output backends:

  1) XLSX (default, no credentials needed):
        python exporter.py country_data.json --xlsx maxwell_sheets_seed.xlsx
     The workbook has one worksheet per country (GB, DE, ES). It can be
     imported straight into Google Sheets (File -> Import -> Upload ->
     "Insert new sheet(s)"), which completes the migration without any API
     credentials.

  2) Google Sheets API (writes directly into the live workbook):
        export GOOGLE_SHEETS_CREDS='{"type":"service_account",...}'
        export SHEET_ID='your-sheet-id'
        python exporter.py country_data.json --sheets
     Requires: pip install google-auth google-api-python-client
     The service account must have edit access to the spreadsheet.

Layout conventions (must stay in lockstep with converter.py):

  * Every section starts with a marker row: "SECTION: <name>" in column A.
  * Flat sections (meta, macro, grid, thesis, pillar/*, technology/*,
    timeSeries/*, playbook) use a FIELD | VALUE header followed by
    key/value rows. All flat values are strings.
  * Table sections (scorecard, loadFactors) use a real header row with one
    column per field, then one row per entry.
  * Scorecard: label, value, unit, meta, trend are REQUIRED columns
    (empty cell -> empty string); help, soWhat, source are OPTIONAL
    (empty cell -> key omitted from the object).
  * Pillar rows are written as row_N_label / row_N_value / row_N_help.
    An absent help cell means the row is a 2-tuple [label, value];
    a populated help cell means a 3-tuple [label, value, help].
  * Technology stats are written as stat_N_label / stat_N_value /
    stat_N_size / stat_N_help. label, value, size are always present
    (size may be the empty string -> 3-tuple keeps its empty third slot);
    help is optional (populated -> 4-tuple).
  * Time series rows are pipe-delimited ("2022|2023|2024"); the literal
    token "null" encodes a JSON null. target2030 rows are comma-delimited
    pairs ("45,47").
  * All cells are written as strings; converter.py applies per-column
    type policy when reading back.
"""

import argparse
import json
import os
import sys

PIPE = "|"


def die(msg: str) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def s(v) -> str:
    """Serialize a scalar to its cell string."""
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def join_pipe(values) -> str:
    cells = [s(v) for v in values]
    for c in cells:
        if PIPE in c:
            die(f"pipe character inside a series value: {c!r}")
    return PIPE.join(cells)


# ---------------------------------------------------------------------------
# Country object -> list of rows (each row is a list of cell strings)
# ---------------------------------------------------------------------------

def country_to_rows(country: dict) -> list[list[str]]:
    rows: list[list[str]] = []

    def marker(name: str):
        rows.append([f"SECTION: {name}"])

    def kv_header():
        rows.append(["FIELD", "VALUE"])

    def blank():
        rows.append([])

    # --- flat object sections -------------------------------------------
    for key in ("meta", "macro"):
        if key in country:
            marker(key)
            kv_header()
            for k, v in country[key].items():
                rows.append([k, s(v)])
            blank()

    # --- single-text sections ---------------------------------------------
    for key in ("grid", "thesis"):
        if key in country:
            marker(key)
            kv_header()
            rows.append(["text", s(country[key])])
            blank()

    # --- scorecard (table) -------------------------------------------------
    if "scorecard" in country:
        marker("scorecard")
        header = ["label", "value", "unit", "meta", "trend", "help", "soWhat", "source"]
        rows.append(header)
        for entry in country["scorecard"]:
            row = []
            for col in header:
                if col in ("help", "soWhat", "source"):
                    row.append(s(entry[col]) if col in entry else "")
                else:
                    row.append(s(entry.get(col, "")))
            rows.append(row)
        blank()

    # --- pillars ------------------------------------------------------------
    for pkey, pillar in country.get("pillars", {}).items():
        marker(f"pillar / {pkey}")
        kv_header()
        for field in ("title", "eyebrow", "accent"):
            rows.append([field, s(pillar[field])])
        for i, r in enumerate(pillar["rows"], start=1):
            rows.append([f"row_{i}_label", s(r[0])])
            rows.append([f"row_{i}_value", s(r[1])])
            if len(r) >= 3:
                if r[2] == "":
                    die(f"pillar {pkey} row {i}: 3-tuple with empty help is not representable")
                rows.append([f"row_{i}_help", s(r[2])])
        blank()

    # --- technologies ---------------------------------------------------------
    for tkey, tech in country.get("technologies", {}).items():
        marker(f"technology / {tkey}")
        kv_header()
        for field in ("name", "tabStat", "headline", "thesis"):
            rows.append([field, s(tech[field])])
        for i, st in enumerate(tech["stats"], start=1):
            rows.append([f"stat_{i}_label", s(st[0])])
            rows.append([f"stat_{i}_value", s(st[1])])
            rows.append([f"stat_{i}_size", s(st[2]) if len(st) >= 3 else ""])
            if len(st) >= 4:
                if st[3] == "":
                    die(f"technology {tkey} stat {i}: 4-tuple with empty help is not representable")
                rows.append([f"stat_{i}_help", s(st[3])])
        blank()

    # --- time series ------------------------------------------------------------
    ts = country.get("timeSeries", {})
    if "capacityHistory" in ts:
        marker("timeSeries / capacityHistory")
        kv_header()
        entries = ts["capacityHistory"]
        keys = []
        for e in entries:
            for k in e:
                if k != "year" and k not in keys:
                    keys.append(k)
        rows.append(["year", join_pipe([e["year"] for e in entries])])
        for k in keys:
            rows.append([k, join_pipe([e.get(k) for e in entries])])
        blank()

    if "target2030" in ts:
        marker("timeSeries / target2030")
        kv_header()
        for k, pair in ts["target2030"].items():
            rows.append([k, ",".join(s(v) for v in pair)])
        blank()

    if "strikeHistory" in ts:
        marker("timeSeries / strikeHistory")
        kv_header()
        entries = ts["strikeHistory"]
        keys = []
        for e in entries:
            for k in e:
                if k != "round" and k not in keys:
                    keys.append(k)
        rows.append(["round", join_pipe([e["round"] for e in entries])])
        for k in keys:
            rows.append([k, join_pipe([e.get(k) for e in entries])])
        blank()

    # --- playbook ------------------------------------------------------------
    if "playbook" in country:
        marker("playbook")
        kv_header()
        for i, play in enumerate(country["playbook"], start=1):
            rows.append([f"play_{i}_title", s(play["title"])])
            rows.append([f"play_{i}_body", s(play["body"])])
        blank()

    # --- loadFactors (table) ------------------------------------------------
    if "loadFactors" in country:
        marker("loadFactors")
        header = ["key", "name", "value", "unit", "meta", "source", "help"]
        rows.append(header)
        for entry in country["loadFactors"]:
            rows.append([s(entry.get(col, "")) for col in header])
        blank()

    return rows


# ---------------------------------------------------------------------------
# Backends
# ---------------------------------------------------------------------------

def write_xlsx(data: dict, path: str) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for iso, country in data.items():
        ws = wb.create_sheet(title=iso)
        for r, row in enumerate(country_to_rows(country), start=1):
            for c, cell in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=cell)
    wb.save(path)
    print(f"Wrote {path} ({', '.join(data.keys())})")


def write_sheets(data: dict) -> None:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build

    creds_json = os.getenv("GOOGLE_SHEETS_CREDS")
    sheet_id = os.getenv("SHEET_ID")
    if not creds_json or not sheet_id:
        die("set GOOGLE_SHEETS_CREDS and SHEET_ID environment variables")

    creds = Credentials.from_service_account_info(
        json.loads(creds_json),
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    service = build("sheets", "v4", credentials=creds)
    meta = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
    existing = {sh["properties"]["title"] for sh in meta["sheets"]}

    requests = []
    for iso in data:
        if iso not in existing:
            requests.append({"addSheet": {"properties": {"title": iso}}})
    if requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=sheet_id, body={"requests": requests}
        ).execute()

    for iso, country in data.items():
        rows = country_to_rows(country)
        service.spreadsheets().values().clear(
            spreadsheetId=sheet_id, range=f"{iso}!A:Z"
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=sheet_id,
            range=f"{iso}!A1",
            valueInputOption="RAW",
            body={"values": rows},
        ).execute()
        print(f"Wrote sheet {iso} ({len(rows)} rows)")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("input", help="country_data.json extracted from the HTML")
    ap.add_argument("--xlsx", metavar="PATH", help="write a local .xlsx workbook")
    ap.add_argument("--sheets", action="store_true", help="write to Google Sheets API")
    args = ap.parse_args()

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    if args.sheets:
        write_sheets(data)
    elif args.xlsx:
        write_xlsx(data, args.xlsx)
    else:
        die("choose an output: --xlsx PATH or --sheets")


if __name__ == "__main__":
    main()
