#!/usr/bin/env python3
"""
converter.py — Part 3 of DATA_PIPELINE_ARCHITECTURE.md (Sheets -> data.json).

Reads the section-based country sheets and produces data.json matching the
Part 2 JSON schema, with a `version` timestamp for cache invalidation.

Two input backends:

  1) Google Sheets API (production; runs inside GitHub Actions):
        export GOOGLE_SHEETS_CREDS='{"type":"service_account",...}'
        export SHEET_ID='your-sheet-id'
        python converter.py --sheets --countries GB DE ES -o data.json
     Requires: pip install google-auth google-api-python-client
     The service account needs read access to the spreadsheet.

  2) Local XLSX (testing / round-trip validation, no credentials):
        python converter.py --xlsx maxwell_sheets_seed.xlsx -o data.json

Layout conventions are documented in exporter.py and must stay in lockstep.

Per-column type policy applied while reading:
  * scorecard.value, loadFactors.value          -> number (fallback: string)
  * timeSeries year                              -> int
  * timeSeries series cells                      -> number; "null"/"" -> null
  * timeSeries strikeHistory "round" row         -> string (never coerced)
  * target2030 cells                              -> numbers (comma-delimited)
  * everything else                               -> string
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone

SECTION_RE = re.compile(r"^SECTION:\s*(.+?)\s*$")
NUM_SUFFIX_RE = re.compile(r"^(row|stat|play)_(\d+)_(\w+)$")


def warn(msg: str) -> None:
    print(f"WARNING: {msg}", file=sys.stderr)


def parse_number(cell: str):
    """Parse a cell into int/float; return the original string on failure."""
    t = cell.strip()
    try:
        return int(t)
    except ValueError:
        pass
    try:
        return float(t)
    except ValueError:
        return cell


def parse_series_cell(cell: str):
    t = cell.strip()
    if t == "" or t.lower() == "null":
        return None
    n = parse_number(t)
    if isinstance(n, str):
        warn(f"non-numeric series value kept as string: {t!r}")
    return n


# ---------------------------------------------------------------------------
# Row stream -> country object
# ---------------------------------------------------------------------------

def split_sections(rows: list[list[str]]) -> list[tuple[str, list[list[str]]]]:
    """Split a sheet's rows into (section_name, rows) chunks."""
    sections: list[tuple[str, list[list[str]]]] = []
    current_name = None
    current_rows: list[list[str]] = []
    for row in rows:
        first = (row[0] if row else "").strip()
        m = SECTION_RE.match(first)
        if m:
            if current_name is not None:
                sections.append((current_name, current_rows))
            current_name = m.group(1)
            current_rows = []
        elif current_name is not None:
            if any(str(c).strip() for c in row):
                current_rows.append([str(c) for c in row])
    if current_name is not None:
        sections.append((current_name, current_rows))
    return sections


def kv_rows(section_rows: list[list[str]]) -> list[tuple[str, str]]:
    """FIELD/VALUE sections: drop the header row, return (field, value) pairs."""
    out = []
    for row in section_rows:
        field = row[0].strip() if len(row) > 0 else ""
        value = row[1] if len(row) > 1 else ""
        if field.upper() == "FIELD":
            continue  # header row
        if field:
            out.append((field, value))
    return out


def numbered_items(pairs: list[tuple[str, str]], prefix: str) -> list[dict]:
    """Collect prefix_N_field pairs into an ordered list of {field: value}."""
    items: dict[int, dict] = {}
    for field, value in pairs:
        m = NUM_SUFFIX_RE.match(field)
        if m and m.group(1) == prefix:
            idx = int(m.group(2))
            items.setdefault(idx, {})[m.group(3)] = value
    return [items[i] for i in sorted(items)]


def table_header(row: list[str]) -> list[str]:
    """Header columns, with trailing empty cells (sheet padding) removed."""
    header = [c.strip() for c in row]
    while header and header[-1] == "":
        header.pop()
    return header


def parse_scorecard(section_rows: list[list[str]]) -> list[dict]:
    if not section_rows:
        return []
    header = table_header(section_rows[0])
    required = ("label", "value", "unit", "meta", "trend")
    optional = ("help", "soWhat", "source")
    entries = []
    for row in section_rows[1:]:
        cells = {header[i]: (row[i] if i < len(row) else "") for i in range(len(header))}
        entry = {}
        for col in required:
            v = cells.get(col, "")
            entry[col] = parse_number(v) if col == "value" else v
        for col in optional:
            v = cells.get(col, "")
            if v != "":
                entry[col] = v
        if entry["label"] == "":
            warn("scorecard row with empty label skipped")
            continue
        entries.append(entry)
    return entries


def parse_load_factors(section_rows: list[list[str]]) -> list[dict]:
    if not section_rows:
        return []
    header = table_header(section_rows[0])
    entries = []
    for row in section_rows[1:]:
        cells = {header[i]: (row[i] if i < len(row) else "") for i in range(len(header))}
        entry = {}
        for col in header:
            v = cells.get(col, "")
            entry[col] = parse_number(v) if col == "value" else v
        if entry.get("key", "") == "":
            warn("loadFactors row with empty key skipped")
            continue
        entries.append(entry)
    return entries


def parse_pillar(pairs: list[tuple[str, str]]) -> dict:
    flat = dict(pairs)
    pillar = {
        "title": flat.get("title", ""),
        "eyebrow": flat.get("eyebrow", ""),
        "accent": flat.get("accent", ""),
        "rows": [],
    }
    for item in numbered_items(pairs, "row"):
        row = [item.get("label", ""), item.get("value", "")]
        if item.get("help", "") != "":
            row.append(item["help"])
        pillar["rows"].append(row)
    return pillar


def parse_technology(pairs: list[tuple[str, str]]) -> dict:
    flat = dict(pairs)
    tech = {
        "name": flat.get("name", ""),
        "tabStat": flat.get("tabStat", ""),
        "headline": flat.get("headline", ""),
        "thesis": flat.get("thesis", ""),
        "stats": [],
    }
    for item in numbered_items(pairs, "stat"):
        stat = [item.get("label", ""), item.get("value", ""), item.get("size", "")]
        if item.get("help", "") != "":
            stat.append(item["help"])
        tech["stats"].append(stat)
    return tech


def parse_series_table(pairs: list[tuple[str, str]], index_field: str, index_type) -> list[dict]:
    """capacityHistory / strikeHistory: pipe-delimited rows -> array of objects."""
    rows = {field: value.split("|") for field, value in pairs}
    if index_field not in rows:
        warn(f"series table missing index row {index_field!r}")
        return []
    index_values = rows.pop(index_field)
    n = len(index_values)
    for field, vals in rows.items():
        if len(vals) != n:
            warn(f"series row {field!r} has {len(vals)} values, expected {n}")
    out = []
    for i in range(n):
        entry = {index_field: index_type(index_values[i].strip())}
        for field, vals in rows.items():
            entry[field] = parse_series_cell(vals[i]) if i < len(vals) else None
        out.append(entry)
    return out


def parse_target2030(pairs: list[tuple[str, str]]) -> dict:
    return {
        field: [parse_number(v.strip()) for v in value.split(",") if v.strip() != ""]
        for field, value in pairs
    }


def parse_playbook(pairs: list[tuple[str, str]]) -> list[dict]:
    return [
        {"title": item.get("title", ""), "body": item.get("body", "")}
        for item in numbered_items(pairs, "play")
    ]


def build_country(rows: list[list[str]], sheet_name: str) -> dict:
    country: dict = {}
    seen_pillars: dict = {}
    seen_techs: dict = {}
    seen_ts: dict = {}

    for name, body in split_sections(rows):
        lowered = name.lower()
        if lowered in ("meta", "macro"):
            country[name] = dict(kv_rows(body))
        elif lowered in ("grid", "thesis"):
            flat = dict(kv_rows(body))
            country[name] = flat.get("text", "")
        elif lowered == "scorecard":
            country["scorecard"] = parse_scorecard(body)
        elif lowered.startswith("pillar"):
            key = name.split("/", 1)[1].strip()
            seen_pillars[key] = parse_pillar(kv_rows(body))
            country.setdefault("pillars", seen_pillars)
        elif lowered.startswith("technology"):
            key = name.split("/", 1)[1].strip()
            seen_techs[key] = parse_technology(kv_rows(body))
            country.setdefault("technologies", seen_techs)
        elif lowered.startswith("timeseries"):
            key = name.split("/", 1)[1].strip()
            pairs = kv_rows(body)
            if key == "capacityHistory":
                seen_ts[key] = parse_series_table(pairs, "year", int)
            elif key == "strikeHistory":
                seen_ts[key] = parse_series_table(pairs, "round", str)
            elif key == "target2030":
                seen_ts[key] = parse_target2030(pairs)
            else:
                warn(f"unknown timeSeries section {key!r} in {sheet_name}; stored as raw pairs")
                seen_ts[key] = dict(pairs)
            country.setdefault("timeSeries", seen_ts)
        elif lowered == "playbook":
            country["playbook"] = parse_playbook(kv_rows(body))
        elif lowered == "loadfactors":
            country["loadFactors"] = parse_load_factors(body)
        else:
            warn(f"unknown section {name!r} in sheet {sheet_name}; skipped")

    # validation
    for required in ("meta", "grid", "thesis", "scorecard", "pillars", "technologies", "timeSeries"):
        if required not in country:
            warn(f"sheet {sheet_name}: missing required section {required!r}")
    if country.get("meta", {}).get("iso", sheet_name) != sheet_name:
        warn(f"sheet {sheet_name}: meta.iso = {country['meta'].get('iso')!r} does not match sheet name")
    return country


# ---------------------------------------------------------------------------
# Backends
# ---------------------------------------------------------------------------

def read_xlsx(path: str, countries: list[str] | None) -> dict[str, list[list[str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    names = countries or wb.sheetnames
    out = {}
    for name in names:
        if name not in wb.sheetnames:
            warn(f"sheet {name!r} not found in {path}")
            continue
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
        out[name] = rows
    return out


def read_sheets(countries: list[str]) -> dict[str, list[list[str]]]:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build

    creds_json = os.getenv("GOOGLE_SHEETS_CREDS")
    sheet_id = os.getenv("SHEET_ID")
    if not creds_json or not sheet_id:
        print("ERROR: set GOOGLE_SHEETS_CREDS and SHEET_ID environment variables", file=sys.stderr)
        sys.exit(1)

    creds = Credentials.from_service_account_info(
        json.loads(creds_json),
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    service = build("sheets", "v4", credentials=creds)
    out = {}
    for name in countries:
        result = service.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range=f"{name}!A:Z",
            valueRenderOption="UNFORMATTED_VALUE",
        ).execute()
        raw = result.get("values", [])
        out[name] = [["" if c is None else str(c) for c in row] for row in raw]
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--sheets", action="store_true", help="read from Google Sheets API")
    src.add_argument("--xlsx", metavar="PATH", help="read from a local .xlsx workbook")
    ap.add_argument("--countries", nargs="*", default=None,
                    help="sheet names to read (default: GB DE ES for --sheets, all sheets for --xlsx)")
    ap.add_argument("-o", "--output", default="data.json")
    args = ap.parse_args()

    if args.sheets:
        countries = args.countries or ["GB", "DE", "ES"]
        sheets = read_sheets(countries)
    else:
        sheets = read_xlsx(args.xlsx, args.countries)

    data = {
        "version": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "countries": {name: build_country(rows, name) for name, rows in sheets.items()},
    }

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"Wrote {args.output} (version {data['version']}, countries: {', '.join(data['countries'])})")


if __name__ == "__main__":
    main()
